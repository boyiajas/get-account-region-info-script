#!/usr/bin/env python3
import argparse
import os
import re
import sys
import time
from typing import Any

from env_config import load_env_file
import requests
from openpyxl import load_workbook

LEGALSUITE_API_BASE = "https://api.legalsuite.net"

SHEET_NAMES = [
    "Collections-UMH",
    "Banking Law Lit- UMH",
    "Conveyancing-UMH",
]

USER_HEADER = "User"
LICENSE_HEADER = "Licenses"


def normalize_name(name: Any) -> str:
    if name is None:
        return ""
    if not isinstance(name, str):
        name = str(name)
    normalized = name.strip().lower()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def post_with_retry(url: str, headers: dict[str, str], params: dict[str, Any] | None = None, timeout: int = 60) -> requests.Response:
    last_exc = None
    for attempt in range(1, 4):
        try:
            response = requests.post(url, headers=headers, params=params, timeout=timeout)
            response.raise_for_status()
            return response
        except requests.RequestException as exc:
            last_exc = exc
            if attempt >= 3:
                raise
            delay = 2 if attempt == 1 else 5
            print(f"Request failed (attempt {attempt}/3): {exc}. Retrying in {delay}s...", file=sys.stderr)
            time.sleep(delay)
    raise last_exc


def get_headers(api_key: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/x-www-form-urlencoded",
    }


def fetch_all_license_rows(api_base: str, api_key: str) -> list[dict[str, Any]]:
    url = f"{api_base.rstrip('/')}/licensed/get"
    print("Fetching all licenses from Legal Suite...")
    response = post_with_retry(url, headers=get_headers(api_key), params=None, timeout=120)
    payload = response.json()
    rows = payload.get("data", [])
    print(f"Received {len(rows)} license rows from /licensed/get")
    return rows


def fetch_license_employees(api_base: str, api_key: str, license_id: str) -> list[dict[str, Any]]:
    url = f"{api_base.rstrip('/')}/licensed/get"
    params = [("where[]", f"Licensed.DocGenID,=,{license_id}")]
    response = post_with_retry(url, headers=get_headers(api_key), params=params, timeout=120)
    payload = response.json()
    items = payload.get("data", [])
    print(f"  - fetched {len(items)} employees for License ID {license_id}")
    return items


def build_license_employee_map(api_base: str, api_key: str) -> dict[str, list[str]]:
    rows = fetch_all_license_rows(api_base, api_key)
    license_names: dict[str, str] = {}
    license_ids: list[str] = []

    for row in rows:
        license_id = str(row.get("docgenrecordid") or row.get("docgenid") or row.get("DocGenID") or "").strip()
        license_name = str(row.get("docgendescription") or row.get("docgencode") or row.get("DocGenDescription") or "").strip()
        if not license_id:
            continue
        if license_name:
            license_names[license_id] = license_name
        if license_id not in license_ids:
            license_ids.append(license_id)

    license_employees: dict[str, list[str]] = {}
    print("Fetching license employees from the second API call for each license...")
    for license_id in sorted(license_ids):
        rows_for_license = fetch_license_employees(api_base, api_key, license_id)
        for row in rows_for_license:
            employee_name = str(row.get("employeename") or row.get("employeeloginid") or "").strip()
            if employee_name:
                license_employees.setdefault(license_id, []).append(employee_name)

    normalized_map: dict[str, list[str]] = {}
    for license_id, names in license_employees.items():
        license_name = license_names.get(license_id, license_id)
        normalized = [normalize_name(name) for name in names if normalize_name(name)]
        normalized_map[license_name] = sorted(set(normalized))

    print(f"Prepared license -> employee map for {len(normalized_map)} licenses")
    return normalized_map


def update_workbook(excel_path: str, license_map: dict[str, list[str]], sheet_names: list[str]) -> None:
    print(f"Loading workbook {excel_path}...")
    workbook = load_workbook(excel_path)
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            print(f"Skipping missing worksheet: {sheet_name}")
            continue

        ws = workbook[sheet_name]
        print(f"Processing worksheet: {sheet_name}")

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if USER_HEADER not in header:
            print(f"  - missing '{USER_HEADER}' header in worksheet '{sheet_name}', skipping")
            continue
        if LICENSE_HEADER not in header:
            license_col = len(header) + 1
            ws.cell(row=1, column=license_col, value=LICENSE_HEADER)
            header.append(LICENSE_HEADER)
            print(f"  - added missing '{LICENSE_HEADER}' header")
        user_col = header.index(USER_HEADER) + 1
        license_col = header.index(LICENSE_HEADER) + 1

        row_count = 0
        updated_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            row_count += 1
            user_cell = row[user_col - 1]
            user_value = user_cell.value
            if user_value is None or str(user_value).strip() == "":
                continue
            normalized_user = normalize_name(user_value)
            if not normalized_user:
                continue

            found_licenses = [license_name for license_name, employees in license_map.items() if normalized_user in employees]
            if found_licenses:
                cell_value = "; ".join(sorted(found_licenses))
            else:
                cell_value = ""

            current_license_value = row[license_col - 1].value
            if current_license_value != cell_value:
                ws.cell(row=user_cell.row, column=license_col, value=cell_value)
                updated_count += 1

            print(
                f"  Row {user_cell.row}: user='{user_value}' -> licenses={len(found_licenses)} "
                f"({'; '.join(found_licenses) if found_licenses else 'none'})"
            )

        print(f"  Completed {row_count} rows, updated {updated_count} cells in '{sheet_name}'.")

    backup_path = excel_path + ".bak"
    if not os.path.exists(backup_path):
        print(f"Saving backup workbook to {backup_path}.")
        with open(excel_path, "rb") as original, open(backup_path, "wb") as backup:
            backup.write(original.read())
    workbook.save(excel_path)
    workbook.close()
    print(f"Workbook saved: {excel_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update LegalSuite licenses for users in an Excel workbook.")
    parser.add_argument(
        "--excel-path",
        default="LegalSuite Training Schedule.xlsx",
        help="Path to the Excel workbook to update.",
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("LEGALSUITE_API_KEY", ""),
        help="LegalSuite API key, or set LEGALSUITE_API_KEY in the environment.",
    )
    parser.add_argument(
        "--api-base",
        default=LEGALSUITE_API_BASE,
        help="LegalSuite API base URL.",
    )
    return parser.parse_args()


def main() -> int:
    load_env_file()
    args = parse_args()
    if not args.api_key:
        print("ERROR: LegalSuite API key missing. Set LEGALSUITE_API_KEY or provide --api-key.", file=sys.stderr)
        return 1
    if not os.path.exists(args.excel_path):
        print(f"ERROR: Excel file not found: {args.excel_path}", file=sys.stderr)
        return 1

    license_map = build_license_employee_map(args.api_base, args.api_key)
    update_workbook(args.excel_path, license_map, SHEET_NAMES)
    return 0


if __name__ == "__main__":
    sys.exit(main())
