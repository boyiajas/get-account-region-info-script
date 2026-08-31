#!/usr/bin/env python3
"""
Process all worksheets in a portfolio workbook, look up account numbers in Legal
Suite, and append matter details to each row in a new workbook.
"""

import os
import sys
import time
from argparse import ArgumentParser
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    from env_config import load_env_file
    load_env_file()
except ImportError:
    env_file = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_file):
        with open(env_file, encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ[key] = value


LEGALSUITE_API_BASE = "https://api.legalsuite.net"
API_KEY_ENVS = [
    "LEGALSUITE_API_KEY",
    "LEGALSUITE_WC_API_KEY",
]
MAX_ATTEMPTS = 3
RETRY_DELAYS = [1, 2, 4]
SOURCE_WORKBOOK = "Strauss Daly Inc Portfolio.xlsx"
OUTPUT_WORKBOOK = "Strauss Daly Inc Portfolio with matter details.xlsx"
MAX_WORKERS = 6
DEFAULT_ACCOUNT_COLUMN_INDEX = 2
START_ROW = 2
FILE_NOTE_LIMIT = 3
LOOKUP_MODE_ACCOUNT_NUMBER = "account_number"
LOOKUP_MODE_FILE_REFERENCE = "file_reference"
PREFERRED_MATTER_TYPE_ID = "4"
DEFENDANT_ROLE_ID = "103"
DEFENDANT_SORTER = "1"
ACCOUNT_HEADER_CANDIDATES = {
    "account_number",
    "account number",
    "accountnumber",
    "account_no",
    "account no",
}
FILE_REFERENCE_HEADER_CANDIDATES = {
    "file_reference",
    "file reference",
    "fileref",
    "file ref",
}
OUTPUT_HEADERS = [
    "file_reference",
    "branch",
    "defendant_first_name",
    "defendant_surname",
    "matter_description",
    "formatteddateinstructed",
    "casenumber",
    "laststagedescription",
    "last_file_note_1",
    "last_file_note_2",
    "last_file_note_3",
]
OUTPUT_COLUMN_WIDTHS = {
    "file_reference": 22,
    "branch": 24,
    "defendant_first_name": 26,
    "defendant_surname": 26,
    "matter_description": 40,
    "formatteddateinstructed": 20,
    "casenumber": 20,
    "laststagedescription": 30,
    "last_file_note_1": 60,
    "last_file_note_2": 60,
    "last_file_note_3": 60,
}


def parse_args() -> tuple[str, Optional[str]]:
    parser = ArgumentParser(
        description=(
            "Process all worksheets in a workbook, look up account numbers in "
            "Legal Suite, and append matter details to each row in a new workbook."
        )
    )
    parser.add_argument(
        "source_workbook",
        nargs="?",
        default=SOURCE_WORKBOOK,
        help=f"Source workbook to process. Defaults to {SOURCE_WORKBOOK!r}.",
    )
    parser.add_argument(
        "-o",
        "--output-workbook",
        dest="output_workbook",
        help="Destination workbook path. Defaults to a derived filename for custom sources.",
    )
    args = parser.parse_args()
    return args.source_workbook, args.output_workbook


def resolve_workbook_path(script_dir: str, workbook_name: str) -> str:
    workbook_path = Path(workbook_name)
    if workbook_path.is_absolute():
        return str(workbook_path)
    return str(Path(script_dir) / workbook_path)


def default_output_path(source_path: str) -> str:
    source = Path(source_path)
    if source.name == SOURCE_WORKBOOK:
        return str(source.with_name(OUTPUT_WORKBOOK))
    return str(source.with_name(f"{source.stem} with matter details{source.suffix}"))


@dataclass
class LegalSuiteClient:
    api_base: str
    api_key: str
    stage_description_cache: dict[str, str] = field(default_factory=dict)

    def _headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }

    def _post(self, endpoint: str, data) -> Optional[list[dict]]:
        url = f"{self.api_base}/{endpoint}"
        response = post_with_retry(url, self._headers(), data, timeout=60)
        if response is None:
            return None

        try:
            payload = response.json()
        except Exception as exc:
            print(f"Error parsing Legal Suite response from {endpoint}: {exc}")
            return None

        return payload.get("data", [])

    def get_matter_by_theirref(self, account_number: str) -> Optional[dict]:
        items = self._post(
            "matter/get",
            {"where[]": f"Matter.TheirRef,=,{account_number}"},
        )
        if not items:
            return None
        return max(items, key=score_matter)

    def get_matter_by_fileref(self, file_reference: str) -> Optional[dict]:
        items = self._post(
            "matter/get",
            {"where[]": f"Matter.FileRef,=,{file_reference}"},
        )
        if not items:
            return None
        return max(items, key=score_matter)

    def get_matparties_for_matter(self, matter_id: str) -> list[dict]:
        items = self._post(
            "matparty/get",
            [
                ("where[]", f"MatParty.MatterID,=,{matter_id}"),
                ("where[]", f"MatParty.Sorter,=,{DEFENDANT_SORTER}"),
                ("where[]", f"MatParty.RoleID,=,{DEFENDANT_ROLE_ID}"),
            ],
        )
        return items or []

    def get_party_by_id(self, party_id: str) -> Optional[dict]:
        items = self._post(
            "party/get",
            {"where[]": f"Party.RecordID,=,{party_id}"},
        )
        if not items:
            return None
        return items[0]

    def get_parlang_by_partyid_and_languageid(
        self,
        party_id: str,
        language_id: str = "1",
    ) -> Optional[dict]:
        items = self._post(
            "parlang/get",
            [
                ("where[]", f"ParLang.PartyID,=,{party_id}"),
                ("where[]", f"ParLang.LanguageID,=,{language_id}"),
            ],
        )
        if not items:
            return None
        return items[0]

    def get_filenotes_for_matter(self, matter_id: str) -> list[dict]:
        items = self._post(
            "filenote/get",
            {"where[]": f"FileNote.MatterID,=,{matter_id}"},
        )
        return items or []

    def get_stage_by_id(self, stage_id: str) -> Optional[dict]:
        items = self._post(
            "stage/get",
            {"where[]": f"Stage.RecordID,=,{stage_id}"},
        )
        if not items:
            return None
        return items[0]

    def get_stage_description(self, stage_id: str) -> str:
        normalized_stage_id = str(stage_id or "").strip()
        if not normalized_stage_id:
            return ""
        cached = self.stage_description_cache.get(normalized_stage_id)
        if cached is not None:
            return cached
        stage = self.get_stage_by_id(normalized_stage_id)
        description = str(stage.get("description") or "").strip() if stage else ""
        self.stage_description_cache[normalized_stage_id] = description
        return description


def post_with_retry(url: str, headers: dict[str, str], data, timeout: int = 60) -> Optional[requests.Response]:
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = requests.post(url, headers=headers, data=data, timeout=timeout)
            if response.status_code >= 500:
                response.raise_for_status()
            return response
        except (requests.Timeout, requests.ConnectionError, requests.HTTPError) as exc:
            if attempt >= MAX_ATTEMPTS:
                print(f"Failed after {MAX_ATTEMPTS} attempts: {exc}")
                return None
            delay = RETRY_DELAYS[min(attempt - 1, len(RETRY_DELAYS) - 1)]
            print(f"Retry attempt {attempt}/{MAX_ATTEMPTS} after {delay}s...")
            time.sleep(delay)
    return None


def get_clients() -> dict[str, LegalSuiteClient]:
    clients = {}
    for env_name in API_KEY_ENVS:
        api_key = os.getenv(env_name, "").strip()
        if api_key:
            clients[env_name] = LegalSuiteClient(LEGALSUITE_API_BASE, api_key)
    return clients


def score_matter(matter: dict) -> tuple[int, int, int]:
    archive_status = str(matter.get("archivestatus") or "").strip()
    is_active = 1 if archive_status in {"", "0"} else 0
    is_preferred_type = 1 if str(matter.get("mattertypeid") or "") == PREFERRED_MATTER_TYPE_ID else 0
    record_id = int(matter.get("recordid") or 0)
    return (is_preferred_type, is_active, record_id)


def parse_int(value: object) -> int:
    try:
        return int(str(value).strip() or "0")
    except (TypeError, ValueError):
        return 0


def split_party_name(full_name: str) -> tuple[str, str]:
    cleaned = " ".join(full_name.replace(",", " , ").split()).strip(" ,")
    if not cleaned:
        return "", ""

    if "," in full_name:
        left, right = [part.strip() for part in full_name.split(",", 1)]
        return right, left

    parts = cleaned.split()
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[1:]), parts[0]


def extract_name_from_description(description: str, account_number: str) -> tuple[str, str]:
    if not description:
        return "", ""

    candidate = description.replace(account_number, " ")
    candidate = candidate.replace("_", " ").replace("/", " ")
    candidate = candidate.replace("-", " ")
    candidate = " ".join(candidate.split())
    if not candidate:
        return "", ""

    return split_party_name(candidate)


def get_defendant_name(
    matter: dict,
    client: LegalSuiteClient,
    reference_value: str,
) -> tuple[str, str]:
    matter_id = str(matter.get("recordid") or "")
    if matter_id:
        matparties = client.get_matparties_for_matter(matter_id)
        if matparties:
            party_id = str(matparties[0].get("partyid") or "")
            if party_id:
                parlang = client.get_parlang_by_partyid_and_languageid(party_id, "1")
                if parlang:
                    parlang_first = (
                        parlang.get("firstname")
                        or parlang.get("firstnames")
                        or parlang.get("forename")
                        or ""
                    ).strip()
                    parlang_surname = (
                        parlang.get("surname")
                        or parlang.get("lastname")
                        or parlang.get("name")
                        or ""
                    ).strip()
                    if parlang_first or parlang_surname:
                        if parlang_first and parlang_surname:
                            return parlang_first, parlang_surname
                        return split_party_name(parlang_surname)

                party = client.get_party_by_id(party_id)
                if party:
                    direct_first = (
                        party.get("firstname")
                        or party.get("firstnames")
                        or party.get("forename")
                        or party.get("contactname")
                        or ""
                    ).strip()
                    direct_surname = (
                        party.get("surname")
                        or party.get("lastname")
                        or ""
                    ).strip()
                    if direct_first or direct_surname:
                        return direct_first, direct_surname
                    return split_party_name(str(party.get("name") or ""))

    return extract_name_from_description(
        str(matter.get("description") or ""),
        reference_value,
    )


def format_filenote(note: dict) -> str:
    description = " ".join(str(note.get("description") or "").split())
    timestamp = (
        str(note.get("formatteddatetime") or "").strip()
        or str(note.get("formatteddate") or "").strip()
    )
    if timestamp and description:
        return f"{timestamp} | {description}"
    return timestamp or description


def get_last_file_notes(
    matter_id: str,
    client: LegalSuiteClient,
    limit: int = FILE_NOTE_LIMIT,
) -> list[str]:
    notes = client.get_filenotes_for_matter(matter_id)
    if not notes:
        return [""] * limit

    latest_notes = sorted(
        notes,
        key=lambda note: (
            parse_int(note.get("date")),
            parse_int(note.get("time")),
            parse_int(note.get("recordid")),
        ),
        reverse=True,
    )[:limit]
    formatted_notes = [format_filenote(note) for note in latest_notes]
    return formatted_notes + [""] * (limit - len(formatted_notes))


def build_output_defaults(fill_value: str) -> dict[str, str]:
    return {header_name: fill_value for header_name in OUTPUT_HEADERS}


def search_matter(
    lookup_mode: str,
    lookup_value: str,
    clients: dict[str, LegalSuiteClient],
) -> tuple[Optional[dict], Optional[LegalSuiteClient]]:
    for env_name in API_KEY_ENVS:
        client = clients.get(env_name)
        if not client:
            continue
        if lookup_mode == LOOKUP_MODE_FILE_REFERENCE:
            matter = client.get_matter_by_fileref(lookup_value)
        else:
            matter = client.get_matter_by_theirref(lookup_value)
        if matter:
            return matter, client
    return None, None


def process_lookup(lookup_mode: str, lookup_value: str, clients: dict[str, LegalSuiteClient]) -> dict[str, str]:
    matter, client = search_matter(lookup_mode, lookup_value, clients)
    if matter and client:
        reference_value = str(matter.get("theirref") or lookup_value)
        defendant_first_name, defendant_surname = get_defendant_name(
            matter,
            client,
            reference_value,
        )
        matter_id = str(matter.get("recordid") or "")
        last_file_notes = get_last_file_notes(matter_id, client) if matter_id else [""] * FILE_NOTE_LIMIT
        last_stage_description = client.get_stage_description(str(matter.get("laststageid") or ""))
        result = {
            "file_reference": matter.get("fileref", ""),
            "branch": matter.get("branchdescription", ""),
            "defendant_first_name": defendant_first_name,
            "defendant_surname": defendant_surname,
            "matter_description": matter.get("description", ""),
            "formatteddateinstructed": matter.get("formatteddateinstructed", "") or "",
            "casenumber": matter.get("casenumber", "") or "",
            "laststagedescription": last_stage_description,
        }
        for index, note_text in enumerate(last_file_notes, start=1):
            result[f"last_file_note_{index}"] = note_text
        return result

    return build_output_defaults("NOT FOUND")


def normalize_account_number(value: object) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    return text if text.isdigit() else ""


def normalize_file_reference(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_value(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def find_column_index(header_row: list[object], candidates: set[str]) -> Optional[int]:
    for index, value in enumerate(header_row, start=1):
        if normalize_header_value(value) in candidates:
            return index
    return None


def find_lookup_strategy(header_row: list[object]) -> tuple[str, int]:
    file_reference_column_index = find_column_index(header_row, FILE_REFERENCE_HEADER_CANDIDATES)
    if file_reference_column_index:
        return LOOKUP_MODE_FILE_REFERENCE, file_reference_column_index

    account_column_index = find_column_index(header_row, ACCOUNT_HEADER_CANDIDATES)
    if account_column_index:
        return LOOKUP_MODE_ACCOUNT_NUMBER, account_column_index

    return LOOKUP_MODE_ACCOUNT_NUMBER, DEFAULT_ACCOUNT_COLUMN_INDEX


def normalize_lookup_value(lookup_mode: str, value: object) -> str:
    if lookup_mode == LOOKUP_MODE_FILE_REFERENCE:
        return normalize_file_reference(value)
    return normalize_account_number(value)


def read_workbook_rows(workbook_path: str) -> tuple[list[dict], list[tuple[str, str]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_payloads = []
    unique_lookup_keys: set[tuple[str, str]] = set()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not rows:
            sheet_payloads.append(
                {
                    "sheet_name": sheet_name,
                    "rows": [],
                    "lookup_mode": LOOKUP_MODE_ACCOUNT_NUMBER,
                    "lookup_column_index": DEFAULT_ACCOUNT_COLUMN_INDEX,
                }
            )
            continue

        lookup_mode, lookup_column_index = find_lookup_strategy(rows[0])
        print(f"Sheet '{sheet_name}': using {lookup_mode} column {lookup_column_index}")

        for row_idx in range(START_ROW - 1, len(rows)):
            row = rows[row_idx]
            lookup_value = normalize_lookup_value(
                lookup_mode,
                row[lookup_column_index - 1] if len(row) >= lookup_column_index else None,
            )
            if lookup_value:
                unique_lookup_keys.add((lookup_mode, lookup_value))

        sheet_payloads.append(
            {
                "sheet_name": sheet_name,
                "rows": rows,
                "lookup_mode": lookup_mode,
                "lookup_column_index": lookup_column_index,
            }
        )

    return sheet_payloads, sorted(unique_lookup_keys)


def build_lookup_results(
    lookup_keys: list[tuple[str, str]],
    clients: dict[str, LegalSuiteClient],
) -> dict[str, dict[str, str]]:
    lookup_results: dict[tuple[str, str], dict[str, str]] = {}
    completed = 0
    found_count = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_lookup = {
            executor.submit(process_lookup, lookup_mode, lookup_value, clients): (lookup_mode, lookup_value)
            for lookup_mode, lookup_value in lookup_keys
        }
        for future in as_completed(future_to_lookup):
            lookup_key = future_to_lookup[future]
            result = future.result()
            lookup_results[lookup_key] = result
            completed += 1
            if result["file_reference"] != "NOT FOUND":
                found_count += 1
            if completed % 50 == 0 or completed == len(lookup_keys):
                print(
                    f"Processed {completed}/{len(lookup_keys)} unique lookups "
                    f"(found {found_count})"
                )

    return lookup_results


def create_excel_report(
    sheet_payloads: list[dict],
    lookup_results: dict[tuple[str, str], dict[str, str]],
    output_path: str,
) -> tuple[int, int]:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    found_rows = 0
    processed_rows = 0

    for sheet_payload in sheet_payloads:
        worksheet = workbook.create_sheet(title=sheet_payload["sheet_name"])
        rows = sheet_payload["rows"]
        if not rows:
            continue

        lookup_mode = sheet_payload["lookup_mode"]
        lookup_column_index = sheet_payload["lookup_column_index"]
        header = list(rows[0]) + OUTPUT_HEADERS
        worksheet.append(header)

        for row_idx, row in enumerate(rows[1:], start=2):
            output_row = list(row)
            lookup_value = normalize_lookup_value(
                lookup_mode,
                row[lookup_column_index - 1] if len(row) >= lookup_column_index else None,
            )
            if row_idx >= START_ROW and lookup_value:
                processed_rows += 1
                result = lookup_results.get((lookup_mode, lookup_value), build_output_defaults("NOT FOUND"))
                if result["file_reference"] != "NOT FOUND":
                    found_rows += 1
            else:
                result = build_output_defaults("")

            output_row.extend(result[header_name] for header_name in OUTPUT_HEADERS)
            worksheet.append(output_row)

        output_column_start = len(rows[0]) + 1
        for offset, header_name in enumerate(OUTPUT_HEADERS):
            column_letter = get_column_letter(output_column_start + offset)
            width = OUTPUT_COLUMN_WIDTHS.get(header_name, 24)
            worksheet.column_dimensions[column_letter].width = width

    workbook.save(output_path)
    return processed_rows, found_rows


def main() -> int:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    source_workbook, output_workbook = parse_args()
    source_path = resolve_workbook_path(script_dir, source_workbook)
    output_path = (
        resolve_workbook_path(script_dir, output_workbook)
        if output_workbook
        else default_output_path(source_path)
    )

    if not os.path.exists(source_path):
        print(f"ERROR: Source workbook not found: {source_path}")
        return 1

    clients = get_clients()
    if not clients:
        print("ERROR: No Legal Suite API keys found in environment")
        return 1

    sheet_payloads, unique_lookup_keys = read_workbook_rows(source_path)
    if not unique_lookup_keys:
        print("ERROR: No valid file references or numeric-only account numbers found")
        return 1

    print(
        f"Loaded {len(sheet_payloads)} sheets; found {len(unique_lookup_keys)} unique lookup values "
        f"from row {START_ROW}"
    )
    lookup_results = build_lookup_results(unique_lookup_keys, clients)
    processed_rows, found_rows = create_excel_report(sheet_payloads, lookup_results, output_path)

    print(f"Saved: {output_path}")
    print(f"Total rows processed: {processed_rows}")
    print(f"Found: {found_rows}")
    print(f"Not found: {processed_rows - found_rows}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
