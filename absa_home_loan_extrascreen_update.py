#!/usr/bin/env python3
import argparse
import datetime as dt
import fnmatch
import ftplib
import json
import os
import re
import shutil
import smtplib
import socket
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to process the Excel files.") from exc


API_BASE = "https://api.legalsuite.net"
EXCEL_BASE = dt.datetime(1899, 12, 30)
LEGALSUITE_OFFSET = 36161
LEGALSUITE_MAX_ATTEMPTS = 3
LEGALSUITE_RETRY_DELAYS = (2, 5)

COMMENTS_SCREEN_ID = 553
PTP_SCREEN_ID = 552
DBN_JHB_REGION = "dbn_jhb"
WC_REGION = "wc"
TARGET_FILEREF_PREFIXES_BY_REGION = {
    DBN_JHB_REGION: ("A0038/", "ABS697/"),
    WC_REGION: ("ABS10/", "ABS34/"),
}
FALLBACK_REGION_BY_REGION = {
    DBN_JHB_REGION: WC_REGION,
}

COMMENTS_SLOT_FIELDS = [
    ("field2", "field3", "field4", "field5"),
    ("field6", "field7", "field8", "field9"),
    ("field10", "field11", "field12", "field13"),
    ("field14", "field15", "field16", "field17"),
    ("field18", "field19", "field20", "field21"),
    ("field22", "field23", "field24", "field25"),
]
COMMENTS_MEMO_MAX_LENGTH = 254

COMPLETION_REPORT_TO = [
    "helpdesk@iconis.co.za",
    "dev@iconis.co.za",
]


@dataclass(frozen=True)
class TargetFile:
    name: str
    kind: str
    region: str
    api_key_env: str
    remote_dir: str
    filename_template: str
    screen_id: int


@dataclass
class VerificationWorkbookState:
    source_path: str
    verification_path: str
    workbook: object
    header_indexes: dict[str, dict[str, int]]


TARGET_FILES = [
    TargetFile(
        name="comments",
        kind="comments",
        region=DBN_JHB_REGION,
        api_key_env="LEGALSUITE_API_KEY",
        remote_dir="ABSA Home Loan Legal/APT TO LSW/DBN-JHB/Comments",
        filename_template="ABSA_Home_Loan_Legal_Comments_DBN_JHB_{date}.xlsx",
        screen_id=COMMENTS_SCREEN_ID,
    ),
    TargetFile(
        name="ptp",
        kind="ptp",
        region=DBN_JHB_REGION,
        api_key_env="LEGALSUITE_API_KEY",
        remote_dir="ABSA Home Loan Legal/APT TO LSW/DBN-JHB/PTP",
        filename_template="ABSA_Home_Loan_Legal_PTP_DBN_JHB_{date}.xlsx",
        screen_id=PTP_SCREEN_ID,
    ),
    TargetFile(
        name="comments_wc",
        kind="comments",
        region=WC_REGION,
        api_key_env="LEGALSUITE_WC_API_KEY",
        remote_dir="ABSA Home Loan Legal/APT TO LSW/WC/Comments",
        filename_template="ABSA_Home_Loan_Legal_Comments_WC_{date}.xlsx",
        screen_id=377,
    ),
    TargetFile(
        name="ptp_wc",
        kind="ptp",
        region=WC_REGION,
        api_key_env="LEGALSUITE_WC_API_KEY",
        remote_dir="ABSA Home Loan Legal/APT TO LSW/WC/PTP",
        filename_template="ABSA_Home_Loan_Legal_PTP_WC_{date}.xlsx",
        screen_id=376,
    ),
]
TARGET_FILE_BY_KIND_AND_REGION = {
    (target.kind, target.region): target
    for target in TARGET_FILES
}


class VerificationWorkbookRecorder:
    def __init__(self, verification_dir: str, path_roots: list[str]) -> None:
        self._verification_dir = os.path.abspath(verification_dir)
        self._path_roots = [os.path.abspath(path) for path in path_roots if path]
        self._states: dict[str, VerificationWorkbookState] = {}

    def record_row(
        self,
        source_path: str,
        row_number: int,
        status: str,
        notes: str,
        get_response: object | None,
        verified_values: dict[str, object] | None = None,
        worksheet_name: str | None = None,
    ) -> str:
        state = self._ensure_state(source_path)
        worksheet = self._resolve_worksheet(state, worksheet_name)
        values = {
            "Verification Status": status,
            "Verification Timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Verification Notes": notes,
            "Verification GET Response": self._serialize_response(get_response),
        }
        if verified_values:
            values.update(verified_values)

        for header_name, value in values.items():
            column_idx = self._ensure_column(state, worksheet.title, header_name)
            worksheet.cell(row=row_number, column=column_idx).value = value

        return state.verification_path

    def finalize(self) -> list[str]:
        saved_paths: list[str] = []
        for source_path, state in list(self._states.items()):
            state.workbook.save(state.verification_path)
            state.workbook.close()
            saved_paths.append(state.verification_path)
            del self._states[source_path]
        return sorted(saved_paths)

    def _ensure_state(self, source_path: str) -> VerificationWorkbookState:
        source_abs = os.path.abspath(source_path)
        state = self._states.get(source_abs)
        if state is not None:
            return state

        verification_path = self._verification_path(source_abs)
        os.makedirs(os.path.dirname(verification_path), exist_ok=True)
        shutil.copy2(source_abs, verification_path)
        workbook = load_workbook(verification_path, read_only=False, data_only=False)
        state = VerificationWorkbookState(
            source_path=source_abs,
            verification_path=verification_path,
            workbook=workbook,
            header_indexes={},
        )
        self._states[source_abs] = state
        return state

    def _verification_path(self, source_path: str) -> str:
        for root in self._path_roots:
            try:
                rel_path = os.path.relpath(source_path, root)
            except ValueError:
                continue
            if rel_path == ".":
                return os.path.join(self._verification_dir, os.path.basename(source_path))
            if not rel_path.startswith(f"..{os.sep}") and rel_path != "..":
                return os.path.join(self._verification_dir, rel_path)
        return os.path.join(self._verification_dir, os.path.basename(source_path))

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""
        return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())

    @staticmethod
    def _serialize_response(response: object | None) -> str:
        if response in (None, ""):
            return ""
        text = json.dumps(response, default=str, ensure_ascii=True)
        if len(text) > 32000:
            return text[:31997] + "..."
        return text

    @staticmethod
    def _resolve_worksheet(state: VerificationWorkbookState, worksheet_name: str | None):
        if worksheet_name and worksheet_name in state.workbook.sheetnames:
            return state.workbook[worksheet_name]
        return state.workbook.active

    def _ensure_column(self, state: VerificationWorkbookState, worksheet_name: str, header_name: str) -> int:
        header_index = state.header_indexes.get(worksheet_name)
        worksheet = state.workbook[worksheet_name]
        if header_index is None:
            header_index = {}
            max_col = worksheet.max_column or 1
            for idx in range(1, max_col + 1):
                key = self._normalize_header(worksheet.cell(row=1, column=idx).value)
                if key and key not in header_index:
                    header_index[key] = idx
            state.header_indexes[worksheet_name] = header_index

        normalized_name = self._normalize_header(header_name)
        existing_idx = header_index.get(normalized_name)
        if existing_idx is not None:
            return existing_idx

        column_idx = (worksheet.max_column or 0) + 1
        worksheet.cell(row=1, column=column_idx).value = header_name
        header_index[normalized_name] = column_idx
        return column_idx


def load_env_file(path: str) -> None:
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            if not key or key in os.environ:
                continue
            value = value.strip()
            if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
                value = value[1:-1]
            os.environ[key] = value


class HTTPResponseWrapper:
    def __init__(self, status_code: int, body: str) -> None:
        self.status_code = status_code
        self.text = body

    def json(self) -> object:
        return json.loads(self.text)

    def raise_for_status(self) -> None:
        if self.status_code >= 400:
            raise HTTPStatusError(self)


class HTTPStatusError(Exception):
    def __init__(self, response: HTTPResponseWrapper) -> None:
        super().__init__(f"HTTP {response.status_code}: {response.text[:200]}")
        self.response = response


class LegalSuiteResponseError(Exception):
    def __init__(self, message: str, response: object | None = None) -> None:
        super().__init__(message)
        self.response = response


def _post_form(url: str, headers: dict[str, str], data: object, timeout: int) -> HTTPResponseWrapper:
    encoded_data = urllib_parse.urlencode(data).encode("utf-8")
    request = urllib_request.Request(url, data=encoded_data, headers=headers, method="POST")
    try:
        with urllib_request.urlopen(request, timeout=timeout) as response:
            body = response.read().decode("utf-8", errors="replace")
            return HTTPResponseWrapper(response.status, body)
    except urllib_error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        return HTTPResponseWrapper(exc.code, body)


def post_with_retry(url: str, headers: dict[str, str], data: object, timeout: int) -> HTTPResponseWrapper:
    last_exc: Exception | None = None
    for attempt in range(1, LEGALSUITE_MAX_ATTEMPTS + 1):
        try:
            response = _post_form(url, headers=headers, data=data, timeout=timeout)
            if response.status_code >= 500:
                response.raise_for_status()
            return response
        except (TimeoutError, socket.timeout, urllib_error.URLError, HTTPStatusError) as exc:
            retryable = isinstance(exc, (TimeoutError, socket.timeout, urllib_error.URLError))
            if isinstance(exc, HTTPStatusError):
                status_code = exc.response.status_code if exc.response is not None else None
                retryable = status_code is not None and status_code >= 500

            last_exc = exc
            if not retryable or attempt >= LEGALSUITE_MAX_ATTEMPTS:
                raise

            delay = LEGALSUITE_RETRY_DELAYS[min(attempt - 1, len(LEGALSUITE_RETRY_DELAYS) - 1)]
            print(
                f"Legal Suite request failed (attempt {attempt}/{LEGALSUITE_MAX_ATTEMPTS}): {exc}. "
                f"Retrying in {delay}s..."
            )
            time.sleep(delay)

    if last_exc:
        raise last_exc
    raise RuntimeError("Legal Suite request failed without an exception.")


class FTPClient:
    def __init__(self, host: str, user: str, password: str, timeout: int) -> None:
        self._host = host
        self._user = user
        self._password = password
        self._timeout = timeout
        self._ftp: ftplib.FTP | None = None

    def connect(self) -> None:
        missing = [
            name
            for name, value in (
                ("FTP_HOST", self._host),
                ("FTP_USER", self._user),
                ("FTP_PASS", self._password),
            )
            if not value
        ]
        if missing:
            raise ValueError(f"Missing FTP credentials: {', '.join(missing)}")
        self._ftp = ftplib.FTP(self._host, timeout=self._timeout)
        self._ftp.login(self._user, self._password)
        self._ftp.set_pasv(True)

    def close(self) -> None:
        if not self._ftp:
            return
        try:
            self._ftp.quit()
        except ftplib.all_errors:
            self._ftp.close()

    def list_dir(self, remote_dir: str) -> list[str] | None:
        ftp = self._require()
        try:
            return ftp.nlst(remote_dir)
        except ftplib.error_perm:
            pass

        current_dir = None
        try:
            current_dir = ftp.pwd()
            ftp.cwd(remote_dir)
            listing = ftp.nlst()
            return [f"{remote_dir}/{name}" for name in listing]
        except ftplib.error_perm:
            return None
        finally:
            if current_dir:
                try:
                    ftp.cwd(current_dir)
                except ftplib.all_errors:
                    pass

    def mdtm_timestamp(self, remote_path: str) -> dt.datetime | None:
        ftp = self._require()
        try:
            response = ftp.sendcmd(f"MDTM {remote_path}")
        except ftplib.all_errors:
            return None
        parts = response.split()
        if len(parts) != 2 or parts[0] != "213":
            return None
        try:
            return dt.datetime.strptime(parts[1], "%Y%m%d%H%M%S")
        except ValueError:
            return None

    def resolve_remote_file(self, remote_dir: str, filename_tmpl: str) -> tuple[str | None, str | None]:
        listing = self.list_dir(remote_dir)
        if listing is None:
            return None, "missing_dir"

        names = [os.path.basename(item) for item in listing]
        if any(char in filename_tmpl for char in "*?["):
            matches = [name for name in names if fnmatch.fnmatch(name, filename_tmpl)]
            if not matches:
                return None, "missing_file"
            if len(matches) == 1:
                return matches[0], None
            newest_name = max(
                matches,
                key=lambda name: self.mdtm_timestamp(f"{remote_dir}/{name}") or dt.datetime.min,
            )
            return newest_name, None

        if filename_tmpl in names:
            return filename_tmpl, None

        ftp = self._require()
        try:
            ftp.size(f"{remote_dir}/{filename_tmpl}")
            return filename_tmpl, None
        except ftplib.all_errors:
            return None, "missing_file"

    def download_file(self, remote_path: str, local_path: str) -> None:
        ftp = self._require()
        with open(local_path, "wb") as handle:
            ftp.retrbinary(f"RETR {remote_path}", handle.write)

    def _require(self) -> ftplib.FTP:
        if not self._ftp:
            raise RuntimeError("FTP client not connected.")
        return self._ftp


class LegalSuiteClient:
    def __init__(self, api_base: str, api_key: str) -> None:
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key

    def get_matter_by_theirref(
        self,
        theirref: str,
        fileref_prefixes: tuple[str, ...],
    ) -> tuple[dict | None, str | None]:
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.TheirRef,=,{theirref}",
        }
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        items = payload.get("data", [])
        if not items:
            return None, "matter_not_found"
        if len(items) == 1:
            return items[0], None

        prefix_matches = [
            item
            for item in items
            if _fileref_matches_prefixes(item.get("fileref"), fileref_prefixes)
        ]
        if not prefix_matches:
            filerefs = ", ".join(sorted(str(item.get("fileref") or "") for item in items))
            print(
                f"Multiple matters found for TheirRef {theirref}, but none matched allowed "
                f"FileRef prefixes {', '.join(fileref_prefixes)}. Found: {filerefs}"
            )
            return None, "fileref_prefix_not_found"

        ranked = sorted(prefix_matches, key=_matter_rank_key, reverse=True)
        chosen = ranked[0]
        print(
            f"Multiple matters found for TheirRef {theirref}; "
            f"using recordid {chosen.get('recordid')} fileref {chosen.get('fileref')} "
            f"after filtering to prefixes {', '.join(fileref_prefixes)}."
        )
        return chosen, None

    def update_matter_extrascreen(self, payload: dict[str, object]) -> dict | str:
        url = f"{self._api_base}/matdocsc/update"
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            parsed = response.json()
        except ValueError:
            return response.text
        if isinstance(parsed, dict) and parsed.get("errors"):
            raise LegalSuiteResponseError(str(parsed.get("errors")), response=parsed)
        return parsed

    def create_matter_extrascreen(self, payload: dict[str, object]) -> dict | str:
        url = f"{self._api_base}/matdocsc/store"
        response = post_with_retry(url, headers=self._headers(), data=payload, timeout=60)
        response.raise_for_status()
        try:
            parsed = response.json()
        except ValueError:
            return response.text
        if isinstance(parsed, dict) and parsed.get("errors"):
            raise LegalSuiteResponseError(str(parsed.get("errors")), response=parsed)
        return parsed

    def get_matter_extrascreen(self, matter_id: int | str, docscreenid: int | str) -> list[dict]:
        url = f"{self._api_base}/matdocsc/get"
        data = [
            ("where[]", f"MatDocSc.MatterID,=,{matter_id}"),
            ("where[]", f"MatDocSc.DocScreenID,=,{docscreenid}"),
        ]
        response = post_with_retry(url, headers=self._headers(), data=data, timeout=60)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", [])

    def _headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


def _matter_rank_key(item: dict[str, object]) -> tuple[int, int]:
    recordid = _to_int(item.get("recordid"))
    matterid = _to_int(item.get("matterid"))
    return (recordid, matterid)


def _fileref_matches_prefixes(value: object, prefixes: tuple[str, ...]) -> bool:
    fileref = str(value or "").strip().upper()
    return any(fileref.startswith(prefix.upper()) for prefix in prefixes)


def _to_int(value: object) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return 0


def normalize_compare_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, (int, float, Decimal)):
        try:
            return format(Decimal(str(value)).normalize(), "f").rstrip("0").rstrip(".") or "0"
        except (InvalidOperation, ValueError):
            return str(value).strip()

    text = str(value).strip()
    if not text:
        return ""

    numeric_text = text.replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", numeric_text):
        try:
            return format(Decimal(numeric_text).normalize(), "f").rstrip("0").rstrip(".") or "0"
        except InvalidOperation:
            pass

    return text


def compare_extrascreen_payload_to_row(payload: dict[str, object], fetched_row: dict) -> list[tuple[str, object, object]]:
    field_names = sorted(key for key in payload if key.startswith("field"))
    mismatches: list[tuple[str, object, object]] = []
    for field_name in field_names:
        sent_value = payload.get(field_name)
        fetched_value = fetched_row.get(field_name)
        if normalize_compare_value(sent_value) != normalize_compare_value(fetched_value):
            mismatches.append((field_name, sent_value, fetched_value))
    return mismatches


def build_extrascreen_verification_values(payload: dict[str, object], fetched_row: dict) -> dict[str, object]:
    verified_values: dict[str, object] = {}
    for field_name in sorted(key for key in payload if key.startswith("field")):
        verified_values[f"Verified {field_name}"] = fetched_row.get(field_name)
    return verified_values


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(char for char in str(value).strip().lower() if char.isalnum())


def normalize_text(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    return text or None


def truncate_text(value: str, max_length: int) -> str:
    if max_length < 0:
        return value
    return value[:max_length]


def normalize_number(value: object) -> int | float | str | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    text = str(value).strip()
    if not text:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return text
    try:
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        return text


def parse_excel_date(value: object) -> dt.datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    if isinstance(value, (int, float)):
        try:
            return EXCEL_BASE + dt.timedelta(days=float(value))
        except (OverflowError, ValueError):
            return None

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("T", " ").replace("/", "-")
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    )
    for fmt in formats:
        try:
            return dt.datetime.strptime(normalized, fmt)
        except ValueError:
            continue
    return None


def format_legalsuite_time(value: object) -> str:
    parsed = parse_excel_date(value)
    if parsed is None:
        return ""
    return parsed.strftime("%H:%M:%S")


def encode_legalsuite_date(value: object) -> int | str:
    parsed = parse_excel_date(value)
    if parsed is None:
        return ""
    excel_serial = (parsed.date() - EXCEL_BASE.date()).days
    return excel_serial + LEGALSUITE_OFFSET


def ensure_local_path(base_dir: str, remote_dir: str, filename: str) -> str:
    local_dir = os.path.join(base_dir, *remote_dir.split("/"))
    os.makedirs(local_dir, exist_ok=True)
    return os.path.join(local_dir, filename)


def build_comments_updates(path: str, screen_id: int, api_key_env: str, region: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)

    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
            header_map = {normalize_header(value): idx for idx, value in enumerate(header_row)}
            required = {
                "accountnumber": "Account Number",
                "commentmemo": "Comment / Memo",
                "numberdialled": "Number Dialled",
                "date": "Date",
                "branchid": "Branch ID",
            }
            missing = [label for key, label in required.items() if key not in header_map]
            if missing:
                raise ValueError(f"{path}: missing required comments columns: {', '.join(missing)}")

            for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                account_number = normalize_text(row[header_map["accountnumber"]])
                if not account_number:
                    continue
                grouped[account_number].append(
                    {
                        "worksheet_name": worksheet.title,
                        "row_number": row_number,
                        "comment": normalize_text(row[header_map["commentmemo"]]) or "",
                        "number_dialled": normalize_text(row[header_map["numberdialled"]]) or "",
                        "date_value": row[header_map["date"]],
                        "date_sort": parse_excel_date(row[header_map["date"]]) or dt.datetime.min,
                        "branch_id": normalize_text(row[header_map["branchid"]]) or "",
                    }
                )
    finally:
        workbook.close()

    updates: list[dict[str, object]] = []
    for account_number, rows in grouped.items():
        newest_rows = sorted(rows, key=lambda item: (item["date_sort"], item["row_number"]), reverse=True)
        latest_rows = newest_rows[: len(COMMENTS_SLOT_FIELDS)]
        ordered_rows = sorted(latest_rows, key=lambda item: (item["date_sort"], item["row_number"]))
        payload: dict[str, object] = {"field1": account_number}
        branch_id = next((row["branch_id"] for row in reversed(ordered_rows) if row["branch_id"]), "")
        for index, slot_fields in enumerate(COMMENTS_SLOT_FIELDS):
            memo_field, dialled_field, date_field, time_field = slot_fields
            if index < len(ordered_rows):
                row = ordered_rows[index]
                payload[memo_field] = truncate_text(row["comment"], COMMENTS_MEMO_MAX_LENGTH)
                payload[dialled_field] = row["number_dialled"]
                payload[date_field] = encode_legalsuite_date(row["date_value"])
                payload[time_field] = format_legalsuite_time(row["date_value"])
            else:
                payload[memo_field] = ""
                payload[dialled_field] = ""
                payload[date_field] = ""
                payload[time_field] = ""
        payload["field26"] = branch_id
        updates.append(
            {
                "account_number": account_number,
                "screen_id": screen_id,
                "api_key_env": api_key_env,
                "region": region,
                "field_payload": payload,
                "source_rows_used": len(ordered_rows),
                "source_row_numbers": [int(row["row_number"]) for row in ordered_rows],
                "worksheet_name": str(ordered_rows[-1]["worksheet_name"]) if ordered_rows else None,
                "verification_row_number": int(ordered_rows[-1]["row_number"]) if ordered_rows else 2,
                "source_file": path,
            }
        )
    return updates


def build_ptp_updates(path: str, screen_id: int, api_key_env: str, region: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)

    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
            header_map = {normalize_header(value): idx for idx, value in enumerate(header_row)}
            required = {
                "accountnumber": "Account Number",
                "ptpcapturedate": "PTP Capture Date",
                "ptpduedate": "PTP Due Date",
                "ptpamount": "PTP Amount",
                "branchid": "Branch ID",
            }
            missing = [label for key, label in required.items() if key not in header_map]
            if missing:
                raise ValueError(f"{path}: missing required PTP columns: {', '.join(missing)}")

            for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                account_number = normalize_text(row[header_map["accountnumber"]])
                if not account_number:
                    continue
                capture_value = row[header_map["ptpcapturedate"]]
                due_value = row[header_map["ptpduedate"]]
                grouped[account_number].append(
                    {
                        "worksheet_name": worksheet.title,
                        "row_number": row_number,
                        "capture_value": capture_value,
                        "due_value": due_value,
                        "amount_value": row[header_map["ptpamount"]],
                        "branch_id": normalize_text(row[header_map["branchid"]]) or "",
                        "sort_date": (
                            parse_excel_date(capture_value)
                            or parse_excel_date(due_value)
                            or dt.datetime.min
                        ),
                    }
                )
    finally:
        workbook.close()

    updates: list[dict[str, object]] = []
    for account_number, rows in grouped.items():
        latest_row = max(rows, key=lambda item: (item["sort_date"], item["row_number"]))
        payload: dict[str, object] = {
            "field1": account_number,
            "field2": encode_legalsuite_date(latest_row["capture_value"]),
            "field3": encode_legalsuite_date(latest_row["due_value"]),
            "field4": normalize_number(latest_row["amount_value"]) or "",
            "field13": normalize_text(latest_row.get("branch_id")) or "",
        }
        updates.append(
            {
                "account_number": account_number,
                "screen_id": screen_id,
                "api_key_env": api_key_env,
                "region": region,
                "field_payload": payload,
                "source_rows_used": 1,
                "source_row_numbers": [int(latest_row["row_number"])],
                "worksheet_name": str(latest_row["worksheet_name"]),
                "verification_row_number": int(latest_row["row_number"]),
                "source_file": path,
            }
        )
    return updates


def build_updates_for_file(
    kind: str,
    path: str,
    screen_id: int,
    api_key_env: str,
    region: str,
) -> list[dict[str, object]]:
    if kind == "comments":
        return build_comments_updates(path, screen_id=screen_id, api_key_env=api_key_env, region=region)
    if kind == "ptp":
        return build_ptp_updates(path, screen_id=screen_id, api_key_env=api_key_env, region=region)
    raise ValueError(f"Unsupported file kind: {kind}")


def matter_lookup_candidates(item: dict[str, object]) -> list[dict[str, object]]:
    kind = str(item.get("kind") or "")
    primary_region = str(item.get("region") or "")
    candidates = [
        {
            "region": primary_region,
            "api_key_env": str(item.get("api_key_env") or ""),
            "screen_id": int(item["screen_id"]),
            "fallback": False,
        }
    ]
    fallback_region = FALLBACK_REGION_BY_REGION.get(primary_region)
    if not fallback_region or not kind:
        return candidates

    fallback_target = TARGET_FILE_BY_KIND_AND_REGION.get((kind, fallback_region))
    if not fallback_target:
        return candidates

    candidates.append(
        {
            "region": fallback_target.region,
            "api_key_env": fallback_target.api_key_env,
            "screen_id": fallback_target.screen_id,
            "fallback": True,
        }
    )
    return candidates


def required_api_envs_for_updates(updates: list[dict[str, object]]) -> list[str]:
    api_envs: set[str] = set()
    for item in updates:
        for candidate in matter_lookup_candidates(item):
            api_key_env = str(candidate.get("api_key_env") or "")
            if api_key_env:
                api_envs.add(api_key_env)
    return sorted(api_envs)


def download_targets(
    ftp_client: FTPClient,
    base_dir: str,
    date_str: str,
    selected_kinds: set[str],
) -> tuple[list[dict[str, object]], list[str]]:
    downloaded: list[dict[str, object]] = []
    problems: list[str] = []

    for target in TARGET_FILES:
        if target.kind not in selected_kinds:
            continue

        filename = target.filename_template.format(date=date_str)
        resolved_name, status = ftp_client.resolve_remote_file(target.remote_dir, filename)
        if not resolved_name:
            if status == "missing_dir":
                problems.append(f"Missing FTP directory: {target.remote_dir}")
            else:
                problems.append(f"Missing FTP file: {target.remote_dir}/{filename}")
            continue

        local_path = ensure_local_path(base_dir, target.remote_dir, resolved_name)
        remote_path = f"{target.remote_dir}/{resolved_name}"
        print(f"Downloading {remote_path}")
        ftp_client.download_file(remote_path, local_path)
        downloaded.append(
            {
                "kind": target.kind,
                "region": target.region,
                "api_key_env": target.api_key_env,
                "screen_id": target.screen_id,
                "target_name": target.name,
                "remote_path": remote_path,
                "local_path": local_path,
            }
        )

    return downloaded, problems


def update_extrascreens(
    clients_by_api_key_env: dict[str, LegalSuiteClient],
    updates: list[dict[str, object]],
    dry_run: bool,
    verify: bool,
    verification_recorder: VerificationWorkbookRecorder | None = None,
) -> dict[str, object]:
    results: list[dict[str, object]] = []
    success_count = 0
    failure_count = 0

    for item in updates:
        account_number = str(item["account_number"])
        screen_id = int(item["screen_id"])
        api_key_env = str(item.get("api_key_env") or "")
        region = str(item.get("region") or "")
        field_payload = dict(item["field_payload"])
        source_file = str(item["source_file"])
        worksheet_name = str(item.get("worksheet_name") or "")
        verification_row_number = int(item.get("verification_row_number") or 2)
        source_row_numbers = [int(row_number) for row_number in item.get("source_row_numbers", [])]
        source_rows_note = (
            f"Source rows used: {', '.join(str(row_number) for row_number in source_row_numbers)}"
            if source_row_numbers
            else ""
        )

        lookup_attempts: list[dict[str, object]] = []
        matter = None
        matter_lookup_reason = "matter_not_found"
        selected_lookup_candidate: dict[str, object] | None = None
        lookup_error: Exception | None = None
        for candidate in matter_lookup_candidates(item):
            candidate_api_key_env = str(candidate["api_key_env"])
            candidate_region = str(candidate["region"])
            candidate_screen_id = int(candidate["screen_id"])
            fileref_prefixes = TARGET_FILEREF_PREFIXES_BY_REGION.get(candidate_region, tuple())
            client = clients_by_api_key_env.get(candidate_api_key_env)
            if client is None:
                lookup_attempts.append(
                    {
                        "api_key_env": candidate_api_key_env,
                        "region": candidate_region,
                        "screen_id": candidate_screen_id,
                        "reason": "missing_api_client",
                    }
                )
                continue

            try:
                matter, candidate_reason = client.get_matter_by_theirref(account_number, fileref_prefixes)
            except Exception as exc:
                lookup_error = exc
                lookup_attempts.append(
                    {
                        "api_key_env": candidate_api_key_env,
                        "region": candidate_region,
                        "screen_id": candidate_screen_id,
                        "reason": "matter_lookup_error",
                        "error": str(exc),
                    }
                )
                break

            candidate_reason = candidate_reason or ""
            lookup_attempts.append(
                {
                    "api_key_env": candidate_api_key_env,
                    "region": candidate_region,
                    "screen_id": candidate_screen_id,
                    "reason": candidate_reason or "matched",
                    "matched": bool(matter),
                }
            )
            if matter:
                selected_lookup_candidate = candidate
                matter_lookup_reason = None
                break
            matter_lookup_reason = candidate_reason or "matter_not_found"
            if candidate_reason != "matter_not_found":
                break

        if lookup_error is not None:
            failure_count += 1
            results.append(
                {
                    "status": "failed",
                    "reason": "matter_lookup_error",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "lookup_attempts": lookup_attempts,
                    "source_file": source_file,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                    "error": str(lookup_error),
                }
            )
            print(f"Matter lookup failed for TheirRef {account_number}: {lookup_error}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Failed",
                    "; ".join(part for part in (f"matter_lookup_error: {lookup_error}", source_rows_note) if part),
                    None,
                    {
                        "Verification Account Number": account_number,
                    },
                    worksheet_name=worksheet_name,
                )
            continue

        if not matter:
            failure_count += 1
            results.append(
                {
                    "status": "failed",
                    "reason": matter_lookup_reason or "matter_not_found",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "lookup_attempts": lookup_attempts,
                    "source_file": source_file,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(
                f"Matter lookup failed for TheirRef {account_number}: "
                f"{matter_lookup_reason or 'matter_not_found'}"
            )
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Failed",
                    "; ".join(part for part in (matter_lookup_reason or "matter_not_found", source_rows_note) if part),
                    None,
                    {
                        "Verification Account Number": account_number,
                    },
                    worksheet_name=worksheet_name,
                )
            continue

        if selected_lookup_candidate is not None:
            resolved_screen_id = int(selected_lookup_candidate["screen_id"])
            resolved_api_key_env = str(selected_lookup_candidate["api_key_env"])
            resolved_region = str(selected_lookup_candidate["region"])
            if (
                resolved_api_key_env != api_key_env
                or resolved_region != region
                or resolved_screen_id != screen_id
            ):
                print(
                    f"Fallback lookup matched TheirRef {account_number} in region {resolved_region}; "
                    f"switching api_key_env={resolved_api_key_env} screen_id={resolved_screen_id}."
                )
            screen_id = resolved_screen_id
            api_key_env = resolved_api_key_env
            region = resolved_region

        client = clients_by_api_key_env.get(api_key_env)
        if client is None:
            failure_count += 1
            results.append(
                {
                    "status": "failed",
                    "reason": "missing_api_client",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "lookup_attempts": lookup_attempts,
                    "source_file": source_file,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(
                f"Missing API client for TheirRef {account_number}: "
                f"api_key_env={api_key_env or '<blank>'}"
            )
            continue

        matter_id = matter.get("recordid") or matter.get("matterid")
        if not matter_id:
            failure_count += 1
            results.append(
                {
                    "status": "failed",
                    "reason": "missing_matter_id",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "source_file": source_file,
                    "matter": matter,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(f"Matter ID missing for TheirRef {account_number}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Failed",
                    "; ".join(part for part in ("missing_matter_id", source_rows_note) if part),
                    matter,
                    {
                        "Verification Account Number": account_number,
                    },
                    worksheet_name=worksheet_name,
                )
            continue

        payload = {
            "matterid": matter_id,
            "docscreenid": screen_id,
            **field_payload,
        }

        if dry_run:
            success_count += 1
            results.append(
                {
                    "status": "dry_run",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "source_file": source_file,
                    "matter_id": matter_id,
                    "payload": payload,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(f"Dry-run update: TheirRef {account_number} -> matterid {matter_id} -> screen {screen_id}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Dry Run",
                    source_rows_note,
                    None,
                    {
                        "Verification Account Number": account_number,
                        "Verification Matter ID": matter_id,
                        "Verification FileRef": matter.get("fileref"),
                    },
                    worksheet_name=worksheet_name,
                )
            continue

        try:
            preexisting_rows = client.get_matter_extrascreen(matter_id, screen_id) if verify else []
            response = client.update_matter_extrascreen(payload)
            if not verify:
                success_count += 1
                results.append(
                    {
                        "status": "updated",
                        "account_number": account_number,
                        "screen_id": screen_id,
                        "api_key_env": api_key_env,
                        "region": region,
                        "source_file": source_file,
                        "matter_id": matter_id,
                        "response": response,
                        "worksheet_name": worksheet_name,
                        "source_row_numbers": source_row_numbers,
                    }
                )
                print(f"Updated TheirRef {account_number} on screen {screen_id} without GET verification")
                if verification_recorder:
                    verification_recorder.record_row(
                        source_file,
                        verification_row_number,
                        "Updated",
                        "; ".join(part for part in ("GET verification skipped", source_rows_note) if part),
                        None,
                        {
                            "Verification Account Number": account_number,
                            "Verification Matter ID": matter_id,
                            "Verification FileRef": matter.get("fileref"),
                        },
                        worksheet_name=worksheet_name,
                    )
                continue

            verification_rows = client.get_matter_extrascreen(matter_id, screen_id)
            if not verification_rows:
                store_response = None
                if not preexisting_rows:
                    print(
                        f"No existing extrascreen row for TheirRef {account_number} on screen {screen_id}; "
                        "trying matdocsc/store"
                    )
                    try:
                        store_response = client.create_matter_extrascreen(payload)
                    except HTTPStatusError as exc:
                        failure_count += 1
                        error_text = exc.response.text if exc.response is not None else str(exc)
                        results.append(
                            {
                                "status": "failed",
                                "reason": "store_http_error",
                                "account_number": account_number,
                                "screen_id": screen_id,
                                "api_key_env": api_key_env,
                                "region": region,
                                "source_file": source_file,
                                "matter_id": matter_id,
                                "response": response,
                                "error": error_text,
                                "worksheet_name": worksheet_name,
                                "source_row_numbers": source_row_numbers,
                            }
                        )
                        print(
                            f"matdocsc/store failed for TheirRef {account_number} on screen {screen_id}: "
                            f"{error_text}"
                        )
                        if verification_recorder:
                            verification_recorder.record_row(
                                source_file,
                                verification_row_number,
                                "Failed",
                                "; ".join(
                                    part
                                    for part in (
                                        f"matdocsc/store failed: {error_text}",
                                        source_rows_note,
                                    )
                                    if part
                                ),
                                None,
                                {
                                    "Verification Account Number": account_number,
                                    "Verification Matter ID": matter_id,
                                    "Verification FileRef": matter.get("fileref"),
                                },
                                worksheet_name=worksheet_name,
                            )
                        continue
                    verification_rows = client.get_matter_extrascreen(matter_id, screen_id)

                if verification_rows:
                    fetched_row = verification_rows[0]
                    mismatches = compare_extrascreen_payload_to_row(field_payload, fetched_row)
                    verification_values = build_extrascreen_verification_values(field_payload, fetched_row)
                    verification_values["Verification Account Number"] = account_number
                    verification_values["Verification Matter ID"] = matter_id
                    verification_values["Verification FileRef"] = matter.get("fileref")
                    verification_values["Verification Store Response"] = (
                        json.dumps(store_response, default=str, ensure_ascii=True)
                        if store_response not in (None, "")
                        else ""
                    )

                    if mismatches:
                        failure_count += 1
                        mismatch_names = ", ".join(field_name for field_name, _, _ in mismatches)
                        notes = "; ".join(
                            part
                            for part in (
                                "matdocsc/store created a row but values still mismatched",
                                f"Mismatched fields: {mismatch_names}",
                                source_rows_note,
                            )
                            if part
                        )
                        results.append(
                            {
                                "status": "failed",
                                "reason": "verification_mismatch_after_store",
                                "account_number": account_number,
                                "screen_id": screen_id,
                                "api_key_env": api_key_env,
                                "region": region,
                                "source_file": source_file,
                                "matter_id": matter_id,
                                "response": response,
                                "store_response": store_response,
                                "verification": fetched_row,
                                "mismatches": [
                                    {"field_name": field_name, "sent_value": sent_value, "fetched_value": fetched_value}
                                    for field_name, sent_value, fetched_value in mismatches
                                ],
                                "worksheet_name": worksheet_name,
                                "source_row_numbers": source_row_numbers,
                            }
                        )
                        print(
                            f"Verification mismatch for TheirRef {account_number} on screen {screen_id} "
                            f"after matdocsc/store: {mismatch_names}"
                        )
                        if verification_recorder:
                            verification_recorder.record_row(
                                source_file,
                                verification_row_number,
                                "Mismatch",
                                notes,
                                fetched_row,
                                verification_values,
                                worksheet_name=worksheet_name,
                            )
                        continue

                    success_count += 1
                    results.append(
                        {
                            "status": "verified",
                            "account_number": account_number,
                            "screen_id": screen_id,
                            "api_key_env": api_key_env,
                            "region": region,
                            "source_file": source_file,
                            "matter_id": matter_id,
                            "response": response,
                            "store_response": store_response,
                            "verification": fetched_row,
                            "worksheet_name": worksheet_name,
                            "source_row_numbers": source_row_numbers,
                            "extrascreen_created": True,
                        }
                    )
                    print(f"Created and verified TheirRef {account_number} on screen {screen_id}")
                    if verification_recorder:
                        verification_recorder.record_row(
                            source_file,
                            verification_row_number,
                            "Verified",
                            "; ".join(
                                part
                                for part in ("Extrascreen row created via matdocsc/store", source_rows_note)
                                if part
                            ),
                            fetched_row,
                            verification_values,
                            worksheet_name=worksheet_name,
                        )
                    continue

                failure_count += 1
                missing_reason = (
                    "extrascreen_row_missing_after_store"
                    if not preexisting_rows
                    else "verification_missing_get_data"
                )
                notes = "; ".join(
                    part
                    for part in (
                        "No extrascreen data returned after update.",
                        "matdocsc/store attempted" if not preexisting_rows else "",
                        f"Pre-update GET rows: {len(preexisting_rows)}",
                        source_rows_note,
                    )
                    if part
                )
                results.append(
                    {
                        "status": "failed",
                        "reason": missing_reason,
                        "account_number": account_number,
                        "screen_id": screen_id,
                        "api_key_env": api_key_env,
                        "region": region,
                        "source_file": source_file,
                        "matter_id": matter_id,
                        "response": response,
                        "store_response": store_response,
                        "pre_verification": preexisting_rows[0] if preexisting_rows else None,
                        "worksheet_name": worksheet_name,
                        "source_row_numbers": source_row_numbers,
                    }
                )
                print(
                    f"Verification failed for TheirRef {account_number} on screen {screen_id}: "
                    f"no GET data returned (pre-update rows: {len(preexisting_rows)})"
                )
                if verification_recorder:
                    verification_recorder.record_row(
                        source_file,
                        verification_row_number,
                        "Missing GET data",
                        notes,
                        verification_rows,
                        {
                            "Verification Account Number": account_number,
                            "Verification Matter ID": matter_id,
                            "Verification FileRef": matter.get("fileref"),
                        },
                        worksheet_name=worksheet_name,
                    )
                continue

            fetched_row = verification_rows[0]
            mismatches = compare_extrascreen_payload_to_row(field_payload, fetched_row)
            verification_values = build_extrascreen_verification_values(field_payload, fetched_row)
            verification_values["Verification Account Number"] = account_number
            verification_values["Verification Matter ID"] = matter_id
            verification_values["Verification FileRef"] = matter.get("fileref")

            if mismatches:
                failure_count += 1
                mismatch_names = ", ".join(field_name for field_name, _, _ in mismatches)
                notes = "; ".join(
                    part
                    for part in (f"Mismatched fields: {mismatch_names}", source_rows_note)
                    if part
                )
                results.append(
                    {
                        "status": "failed",
                        "reason": "verification_mismatch",
                        "account_number": account_number,
                        "screen_id": screen_id,
                        "source_file": source_file,
                        "matter_id": matter_id,
                        "response": response,
                        "verification": fetched_row,
                        "mismatches": [
                            {"field_name": field_name, "sent_value": sent_value, "fetched_value": fetched_value}
                            for field_name, sent_value, fetched_value in mismatches
                        ],
                        "worksheet_name": worksheet_name,
                        "source_row_numbers": source_row_numbers,
                    }
                )
                print(
                    f"Verification mismatch for TheirRef {account_number} on screen {screen_id}: "
                    f"{mismatch_names}"
                )
                if verification_recorder:
                    verification_recorder.record_row(
                        source_file,
                        verification_row_number,
                        "Mismatch",
                        notes,
                        fetched_row,
                        verification_values,
                        worksheet_name=worksheet_name,
                    )
                continue

            success_count += 1
            results.append(
                {
                    "status": "verified",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "source_file": source_file,
                    "matter_id": matter_id,
                    "response": response,
                    "verification": fetched_row,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(f"Verified TheirRef {account_number} on screen {screen_id}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Verified",
                    source_rows_note,
                    fetched_row,
                    verification_values,
                    worksheet_name=worksheet_name,
                )
        except (HTTPStatusError, LegalSuiteResponseError) as exc:
            failure_count += 1
            if isinstance(exc, HTTPStatusError):
                error_text = exc.response.text if exc.response is not None else str(exc)
            else:
                error_text = str(exc)
            results.append(
                {
                    "status": "failed",
                    "reason": "http_error",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "source_file": source_file,
                    "matter_id": matter_id,
                    "error": error_text,
                    "payload": payload,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(f"Update failed for TheirRef {account_number}: {error_text}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Failed",
                    "; ".join(part for part in (error_text, source_rows_note) if part),
                    None,
                    {
                        "Verification Account Number": account_number,
                        "Verification Matter ID": matter_id,
                        "Verification FileRef": matter.get("fileref"),
                    },
                    worksheet_name=worksheet_name,
                )
        except Exception as exc:  # pragma: no cover
            failure_count += 1
            results.append(
                {
                    "status": "failed",
                    "reason": "unexpected_error",
                    "account_number": account_number,
                    "screen_id": screen_id,
                    "api_key_env": api_key_env,
                    "region": region,
                    "source_file": source_file,
                    "matter_id": matter_id,
                    "error": str(exc),
                    "payload": payload,
                    "worksheet_name": worksheet_name,
                    "source_row_numbers": source_row_numbers,
                }
            )
            print(f"Unexpected failure for TheirRef {account_number}: {exc}")
            if verification_recorder:
                verification_recorder.record_row(
                    source_file,
                    verification_row_number,
                    "Failed",
                    "; ".join(part for part in (str(exc), source_rows_note) if part),
                    None,
                    {
                        "Verification Account Number": account_number,
                        "Verification Matter ID": matter_id,
                        "Verification FileRef": matter.get("fileref"),
                    },
                    worksheet_name=worksheet_name,
                )

    return {
        "results": results,
        "success_count": success_count,
        "failure_count": failure_count,
    }


def normalize_cli_args(argv: list[str]) -> list[str]:
    normalized: list[str] = []
    for arg in argv:
        match = re.fullmatch(r"--day-(\d+)", arg)
        if match:
            normalized.extend(["--days-ago", match.group(1)])
            continue
        normalized.append(arg)
    return normalized


def parse_args(argv: list[str]) -> argparse.Namespace:
    argv = normalize_cli_args(argv)
    parser = argparse.ArgumentParser(
        description="Download today's ABSA Home Loan Legal files from FTP and update Legal Suite extra screens."
    )
    parser.add_argument(
        "--date",
        help="Date in YYYYMMDD format. Defaults to today.",
    )
    parser.add_argument(
        "--days-ago",
        type=int,
        default=0,
        help="Process files from N days ago. Shorthand like --day-1 is also supported.",
    )
    parser.add_argument(
        "--download-dir",
        default="downloads_absa",
        help="Local base directory for FTP downloads.",
    )
    parser.add_argument(
        "--verification-dir",
        default="verification_absa",
        help="Local base directory for verification workbooks.",
    )
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Path to the environment file with FTP and Legal Suite credentials.",
    )
    parser.add_argument(
        "--api-base",
        default=API_BASE,
        help="Legal Suite API base URL.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="FTP connection timeout in seconds.",
    )
    parser.add_argument(
        "--report-json",
        help="Optional path for a JSON report file.",
    )
    parser.add_argument(
        "--only",
        choices=("all", "comments", "ptp"),
        default="all",
        help="Process only one file type.",
    )
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Skip the FTP download step and process already-downloaded files in --download-dir.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Build payloads and fetch matters without updating Legal Suite.",
    )
    parser.add_argument(
        "--no-verify",
        action="store_true",
        help="Skip the Legal Suite GET verification after each update.",
    )
    return parser.parse_args(argv)


def resolve_date(date_arg: str | None, days_ago: int) -> str:
    if not date_arg:
        if days_ago < 0:
            raise SystemExit("days_ago must be 0 or greater.")
        return (dt.datetime.now() - dt.timedelta(days=days_ago)).strftime("%Y%m%d")
    try:
        return dt.datetime.strptime(date_arg, "%Y%m%d").strftime("%Y%m%d")
    except ValueError as exc:
        raise SystemExit("Date must be in YYYYMMDD format.") from exc


def selected_kinds_from_arg(value: str) -> set[str]:
    if value == "all":
        return {"comments", "ptp"}
    return {value}


def existing_local_downloads(base_dir: str, date_str: str, selected_kinds: set[str]) -> tuple[list[dict[str, object]], list[str]]:
    found: list[dict[str, object]] = []
    problems: list[str] = []

    for target in TARGET_FILES:
        if target.kind not in selected_kinds:
            continue
        filename = target.filename_template.format(date=date_str)
        local_path = ensure_local_path(base_dir, target.remote_dir, filename)
        if not os.path.exists(local_path):
            problems.append(f"Missing local file: {local_path}")
            continue
        found.append(
            {
                "kind": target.kind,
                "region": target.region,
                "api_key_env": target.api_key_env,
                "screen_id": target.screen_id,
                "target_name": target.name,
                "remote_path": f"{target.remote_dir}/{filename}",
                "local_path": local_path,
            }
        )

    return found, problems


def build_report_path(download_dir: str, date_str: str) -> str:
    report_dir = os.path.join(download_dir, "_reports")
    os.makedirs(report_dir, exist_ok=True)
    return os.path.join(report_dir, f"absa_home_loan_extrascreen_update_{date_str}.json")


def send_completion_report_email(
    report_path: str,
    report: dict[str, object],
    report_lines: list[str],
) -> bool:
    smtp_host = os.getenv("MAIL_HOST", os.getenv("SMTP_HOST", "")).strip()
    smtp_port = int(os.getenv("MAIL_PORT", os.getenv("SMTP_PORT", "587")).strip() or "587")
    smtp_user = os.getenv("MAIL_USERNAME", os.getenv("SMTP_USER", "")).strip()
    smtp_pass = os.getenv("MAIL_PASSWORD", os.getenv("SMTP_PASS", "")).strip()
    smtp_from = os.getenv("MAIL_FROM_ADDRESS", os.getenv("SMTP_FROM", smtp_user)).strip()
    auth_mode = os.getenv("MAIL_AUTH_MODE", os.getenv("SMTP_AUTH_MODE", "login")).strip().lower()
    smtp_use_auth = auth_mode not in {"none", "noauth", "false", "0", "no"}
    encryption = os.getenv("MAIL_ENCRYPTION", "").strip().lower()
    smtp_use_tls = encryption not in {"", "null", "none", "false", "0", "no"} if "MAIL_ENCRYPTION" in os.environ else (
        os.getenv("SMTP_USE_TLS", "true").strip().lower() not in {"0", "false", "no"}
    )

    missing = [name for name, value in (("SMTP_HOST", smtp_host), ("SMTP_FROM", smtp_from)) if not value]
    if missing:
        message = f"Completion report email skipped: missing SMTP settings: {', '.join(missing)}"
        report_lines.append(message)
        print(message)
        return False

    message = EmailMessage()
    subject_timestamp = dt.datetime.now().strftime("%Y/%m/%d %-H:%M")
    message["Subject"] = f"ABSA Home Loan Feedback Update Report -- {subject_timestamp}"
    message["From"] = smtp_from
    message["To"] = ", ".join(COMPLETION_REPORT_TO)
    message.set_content(
        "Good Day,\n\n"
        "Please find attached the ABSA Home Loan feedback update report for the selected run.\n\n"
        f"Date: {report.get('date', '')}\n"
        f"Prepared updates: {report.get('prepared_updates', 0)}\n"
        f"Success count: {report.get('success_count', 0)}\n"
        f"Failure count: {report.get('failure_count', 0)}\n"
        f"Verification workbooks: {len(report.get('verification_files', []))}\n\n"
        "Kind Regards,\n"
    )

    with open(report_path, "rb") as handle:
        message.add_attachment(
            handle.read(),
            maintype="application",
            subtype="json",
            filename=os.path.basename(report_path),
        )

    last_exc: Exception | None = None
    for attempt in range(1, 4):
        try:
            if attempt > 1:
                print(f"Retrying completion report email (attempt {attempt}/3)...")
            with smtplib.SMTP(smtp_host, smtp_port, timeout=60) as server:
                if smtp_use_tls:
                    server.starttls()
                if smtp_use_auth and smtp_user:
                    server.login(smtp_user, smtp_pass)
                server.send_message(
                    message,
                    from_addr=smtp_from,
                    to_addrs=COMPLETION_REPORT_TO,
                )
            last_exc = None
            break
        except Exception as exc:
            last_exc = exc
            report_lines.append(f"Completion report email attempt {attempt}/3 failed: {exc}")
            print(f"Completion report email attempt {attempt}/3 failed: {exc}", file=sys.stderr)

    if last_exc is not None:
        report_lines.append(f"Completion report email failed: {last_exc}")
        return False

    report_lines.append("Completion report email sent: to=helpdesk@iconis.co.za, dev@iconis.co.za")
    print("Completion report email sent: to=helpdesk@iconis.co.za, dev@iconis.co.za")
    return True


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    load_env_file(args.env_file)

    date_str = resolve_date(args.date, args.days_ago)
    selected_kinds = selected_kinds_from_arg(args.only)

    ftp_host = os.getenv("FTP_HOST", "")
    ftp_user = os.getenv("FTP_USER", "")
    ftp_pass = os.getenv("FTP_PASS", "")

    download_records: list[dict[str, object]]
    download_problems: list[str]

    if args.skip_download:
        download_records, download_problems = existing_local_downloads(args.download_dir, date_str, selected_kinds)
    else:
        ftp_client = FTPClient(ftp_host, ftp_user, ftp_pass, args.timeout)
        ftp_client.connect()
        try:
            download_records, download_problems = download_targets(
                ftp_client=ftp_client,
                base_dir=args.download_dir,
                date_str=date_str,
                selected_kinds=selected_kinds,
            )
        finally:
            ftp_client.close()

    for problem in download_problems:
        print(problem)

    all_updates: list[dict[str, object]] = []
    for download_record in download_records:
        path = str(download_record["local_path"])
        kind = str(download_record["kind"])
        screen_id = int(download_record["screen_id"])
        api_key_env = str(download_record["api_key_env"])
        region = str(download_record["region"])
        print(f"Processing {path}")
        updates = build_updates_for_file(
            kind,
            path,
            screen_id=screen_id,
            api_key_env=api_key_env,
            region=region,
        )
        for item in updates:
            item["kind"] = kind
        all_updates.extend(updates)
        print(
            f"Prepared {len(updates)} {kind} update(s) from {os.path.basename(path)} "
            f"| region={region} | screen_id={screen_id}"
        )

    api_envs_in_run = required_api_envs_for_updates(all_updates)
    missing_api_envs = [env_name for env_name in api_envs_in_run if not os.getenv(env_name, "").strip()]
    if missing_api_envs:
        raise SystemExit(f"Missing Legal Suite API key(s): {', '.join(missing_api_envs)}")
    clients_by_api_key_env = {
        env_name: LegalSuiteClient(args.api_base, os.getenv(env_name, ""))
        for env_name in api_envs_in_run
    }
    verification_recorder = VerificationWorkbookRecorder(
        args.verification_dir,
        path_roots=[args.download_dir],
    )
    try:
        update_summary = update_extrascreens(
            clients_by_api_key_env=clients_by_api_key_env,
            updates=all_updates,
            dry_run=args.dry_run,
            verify=not args.no_verify,
            verification_recorder=verification_recorder,
        )
    finally:
        verification_paths = verification_recorder.finalize()

    report_lines: list[str] = []
    report = {
        "date": date_str,
        "download_dir": os.path.abspath(args.download_dir),
        "verification_dir": os.path.abspath(args.verification_dir),
        "dry_run": args.dry_run,
        "download_problems": download_problems,
        "downloaded_files": download_records,
        "prepared_updates": len(all_updates),
        "success_count": update_summary["success_count"],
        "failure_count": update_summary["failure_count"],
        "verification_files": verification_paths,
        "results": update_summary["results"],
    }

    report_path = args.report_json or build_report_path(args.download_dir, date_str)
    with open(report_path, "w", encoding="utf-8") as handle:
        json.dump(report, handle, indent=2, default=str)

    send_completion_report_email(report_path, report, report_lines)

    print(
        f"Completed with {update_summary['success_count']} success(es) and "
        f"{update_summary['failure_count']} failure(s). Report: {report_path}"
    )
    if verification_paths:
        print(
            f"Verification workbook(s): {len(verification_paths)} saved under "
            f"{os.path.abspath(args.verification_dir)}"
        )
    return 0 if update_summary["failure_count"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
